#!/usr/bin/env python3
# -*- mode: python; tab-width: 4; indent-tabs-mode: nil -*- for emacs

import argparse
import csv
import datetime
import hashlib
import json
from operator import itemgetter
import os
import re
import sys
import time
import warnings

import boto3
import openpyxl
import urllib3

# WARN_URL  = 'https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report.xlsx'
WARN_URL  = 'https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report1.xlsx'
XLSX_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def parse_options():
    parser = argparse.ArgumentParser(
        description="""
Reading and processing the California WARN act (Excel) spreadsheet as it can be found here:

https://edd.ca.gov/en/jobs_and_training/Layoff_Services_WARN
"""
    )
    parser.add_argument('-d', '--debug',
                        help="Enable debug output",
                        default=False,
                        action='store_true')
    parser.add_argument('-v', '--verbose',
                        help="Enable verbose output",
                        default=False,
                        action='store_true')
    parser.add_argument('-n', '--dryrun',
                        help="Request dryrun (noop) mode",
                        default=False,
                        action='store_true')
    parser.add_argument("--summary",
                        help="Specify an alternative name for 'summary.csv'",
                        default="summary.csv")
    parser.add_argument("--excel",
                        help="Specify an alternative name for 'warn_report.xlsx'",
                        default="warn_report.xlsx")
    parser.add_argument("--server",
                        help="Specify an alternative Mastodon server",
                        default="botsin.space")
    parser.add_argument("--token",
                        help="Specify the authorization token. Required for --post option.")
    parser.add_argument("--post",
                        help="Post this to Mastodon? Only use with --update.",
                        default=False,
                        action='store_true')
    parser.add_argument("--sqs",
                        help="Specify the SQS queue to which to post the updates, if any.")

    # The possible actions:
    parser.add_argument('--dump',
                        help="Dump a summary of the entries in the WARN act spreadsheet",
                        default=False,
                        action='store_true')
    parser.add_argument('--fetch',
                        help="Fetch the latest version of the WARN act spreadsheet from its source",
                        default=False,
                        action='store_true')
    parser.add_argument('--update',
                        help="Update the extracted data with this WARN act spreadsheet",
                        default=False,
                        action="store_true")
    parser.add_argument('--search',
                        help="Search entries matching a company in the summary.csv file")

    opts = parser.parse_args()
    excl = 0
    if opts.dump:
        excl += 1
    if opts.fetch:
        excl += 1
    if opts.search:
        excl += 1
    if opts.update:
        excl += 1
    if excl > 1:
        print("The options --dump, --fetch, --search, and --update are mutually exclusive.")
        sys.exit(1)
    if excl == 0:
        opts.dump = True
    if opts.post and not opts.update:
        print("The option --post can only be used in combination with --update.")
        sys.exit(1)
    if opts.post and not opts.token:
        print("The option --post requires that you also use the --token option.")
        sys.exit(1)
    return opts


def load_report(opts):
    with warnings.catch_warnings():
        # This is just to suppress this (irrelevant to me) warnings
        # .../openpyxl/worksheet/_reader.py:329:
        #     UserWarning: Data Validation extension is not supported and will be removed
        warnings.filterwarnings("ignore", category=UserWarning)
        workbook = openpyxl.load_workbook(filename=opts.excel, data_only=True)

    # Get All Sheets
    a_sheet_names = workbook.sheetnames

    # Look for "Detailed WARN Report" sheet (actual sheet has an extra space).
    # If none found, use the first sheet and hope for the best.
    offset = 0
    sheet_name = a_sheet_names[0]
    for name in a_sheet_names:
        if re.match(r"(?i)detailed warn report", name):
            sheet_name = name
            offset = 1
    if opts.debug:
        print(f"Using sheet name <{sheet_name}>")

    o_sheet = workbook[sheet_name]
    headers = {}
    for col in range(o_sheet.max_column):
        o_cell = o_sheet.cell(row=1+offset, column=col+1)
        header = o_cell.value
        if header is None or len(header) == 0:
            # Break out of the loop if we get a header without content
            # or empty content
            break
        header = header.replace("\n", " ")
        header = header.replace("/ ", "/")
        header = header.replace("  ", " ")
        headers[header] = col + 1
        useful_columns = col + 1
    if opts.debug:
        print(json.dumps(headers))

    # Sanity check a few of the expected headers:
    if "No. Of Employees" not in headers:
        print("Missing No. Of Employees column, aborting.")
        sys.exit(1)
    if "Notice Date" not in headers:
        print("Missing Notice Date column, aborting.")
        sys.exit(1)

    return o_sheet, headers, offset, useful_columns


def update_row_summary(row,
                       last_company,
                       last_companies, last_counties, last_addresses, last_effective,
                       company_col, county_col, address_col, effective_col):
    if len(last_counties) > 1:
        row[county_col] = ", ".join(last_counties.keys())
        row[address_col] = f"[Multiple ({len(last_addresses)})]"
    else:
        if len(last_addresses) > 1:
            row[address_col] = f"[Multiple ({len(last_addresses)})]"
    if len(last_effective) > 1:
        dates = list(last_effective.keys())
        dates.sort()
        row[effective_col] = f"{dates[0]} - {dates[-1]}"
    if len(last_companies) > 1:
        row[company_col] = f"{last_company} [Multiple variations ({len(last_companies)})]"


# Sort (first by Notice Date and then by Company name) and group the
# rows by company (or variations thereof), grouping things like
# multiple different addresses, counties/parishes, and effective
# dates, # while showing the total number of employees affected.
#
# The company name simpification is to handle things like:
#
#   "<name> - <#> Building"
#   "<name> - Layoff <#>"
#
# These may or may not be legally distinct units, but they're typically
# clearly related.
def group_entries(rows, csv_headers):
    output_list = []

    headers = {}
    for col, header in enumerate(csv_headers):
        header = header.replace("  ", " ")
        headers[header] = col

    notice_col    = headers["Notice Date"]
    company_col   = headers["Company"]
    county_col    = headers["County/Parish"]
    processed_col = headers['Processed Date']
    effective_col = headers['Effective Date']
    layoff_col    = headers["Layoff/Closure"]
    address_col   = headers['Address']
    employees_col = headers["No. Of Employees"]

    last_row = {}
    last_company = None
    last_companies = {}
    last_counties = {}
    last_addresses = {}
    last_effective = {}
    for row in sorted(rows, key=itemgetter(notice_col, company_col)):
        eff_company = row[company_col]
        match_google = re.match(r"^Google US-(.*)", eff_company)
        if match_google:
            eff_company = "Google US"
        match_company = re.match(r"^(.*) - (.*)$", eff_company)
        if match_company:
            eff_company = match_company.group(1)

        if len(last_row) > 0 and (row[notice_col]     != last_row[notice_col] or
                                  eff_company.lower() != last_company.lower() or
                                  row[processed_col]  != last_row[processed_col] or
                                  row[layoff_col]     != last_row[layoff_col]):
            # We cannot combine this row into the previous one
            update_row_summary(last_row,
                               last_company,
                               last_companies, last_counties, last_addresses, last_effective,
                               company_col, county_col, address_col, effective_col)

            output_list.append(last_row)
            last_row = {}
            # Fall through

        if len(last_row) == 0:
            # Either because this is the first row, or we cannot combine
            # this row into the previous one
            last_row       = row[:] # Copy the content, not the reference
            last_company   = eff_company
            last_companies = {last_row[company_col]:   True}
            last_counties  = {last_row[county_col]:    True}
            last_addresses = {last_row[address_col]:   True}
            last_effective = {last_row[effective_col]: True}
            continue

        # Fold row into last_row
        last_row[employees_col] = int(last_row[employees_col]) + int(row[employees_col])
        last_companies[row[company_col]] = True
        last_counties[row[county_col]] = True
        last_addresses[row[address_col]] = True
        last_effective[row[effective_col]] = True

    if last_row:
        update_row_summary(last_row,
                           last_company,
                           last_companies, last_counties, last_addresses, last_effective,
                           company_col, county_col, address_col, effective_col)
        output_list.append(last_row)

    return output_list


def dump_entries(rows, csv_headers, align=True):
    output_list = []
    output = ""
    headers = {}
    for col, header in enumerate(csv_headers):
        headers[header] = col
    notice_col = headers["Notice Date"]
    last_notice = None
    for row in group_entries(rows, csv_headers):
        notice = row[notice_col]
        if not last_notice or last_notice != notice:
            output += f"NOTICE DATE: {notice}\n\n"
            last_notice = notice
        for col, header in enumerate(csv_headers):
            if col == notice_col:
                continue
            value  = row[col]
            header = header.replace("\n", " ")
            header = header.replace("/ ", "/")
            header = header.replace("  ", " ")
            if align:
                output += f"  {header:16s} : {value}\n"
            else:
                output += f"{header}: {value}\n"
        output_list.append(output)
        output = ""
    return output_list


def do_dump(o_sheet, headers, offset):
    counties = {}
    companies = {}
    for row in range(o_sheet.max_row - 1 - offset):
        county = o_sheet.cell(row=row+2+offset, column=headers["County/Parish"]).value
        if not county or county == "Report Summary":
            break
        if county not in counties:
            counties[county] = 0
        counties[county] += 1
        company = o_sheet.cell(row=row+2+offset, column=headers["Company"]).value
        if company not in companies:
            companies[company] = {}
        action = o_sheet.cell(row=row+2+offset, column=headers["Layoff/Closure"]).value
        employees = o_sheet.cell(row=row+2+offset, column=headers["No. Of Employees"]).value
        if action not in companies[company]:
            companies[company][action] = 0
        companies[company][action] += employees
    print(json.dumps(counties))
    print(json.dumps(companies))


def do_fetch(opts):
    http = urllib3.PoolManager()
    result = http.request('GET', WARN_URL)
    if result.status != 200:
        print(f"Unexpected HTTP status code: {result.status}")
        sys.exit(1)
    if result.headers['content-type'] != XLSX_TYPE:
        print(f"Unexpect content received: {result.headers['content-type']}; {result.encoding}")
        sys.exit(1)
    fname = opts.excel
    tmp_fname = f"{fname}.{os.getpid()}"
    if opts.debug:
        print(f"Creating temporary file {tmp_fname}.")
    with open(tmp_fname, 'wb') as excel:
        excel.write(result.data)
        excel.close()
    if opts.debug:
        print(f"Renaming {tmp_fname} to {fname}")
    os.rename(tmp_fname, fname)


def do_search(opts):
    fname = opts.summary
    csv_headers = None
    rows = []

    try:
        with open(fname, newline='', encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if csv_headers is None:
                    csv_headers = row
                else:
                    rows.append(row)
    except IOError:
        print(f"File {csv} does not exist.")
        sys.exit(1)

    headers = {}
    for col, header in enumerate(csv_headers):
        headers[header] = col
    if opts.debug:
        print(f"Searching for {opts.search} among {len(rows)} rows of data.")
    rows_found = []
    for row in rows:
        company = row[headers["Company"]]
        if re.match(opts.search, company):
            rows_found.append(row)
    if len(rows_found) > 0:
        print("\n".join(dump_entries(rows_found, csv_headers)))
    else:
        print("No matching companies found.")


def send_to_sqs(opts, output_list, list_size):
    # Reenable the event source mapping, first:
    aws_lambda = boto3.client("lambda")
    esm_uuid = os.environ['ESM_UUID']
    result = aws_lambda.update_event_source_mapping(
        UUID=esm_uuid,
        Enabled=True
    )
    print(f"result = {result}")

    sqs = boto3.resource('sqs')
    queue = sqs.Queue(opts.sqs)
    queue.send_message(
        MessageBody=json.dumps(output_list),
        MessageAttributes={
            'index': {
                'DataType': 'Number',
                'StringValue': '1'
            },
            'sqs_url': {
                'DataType': 'String',
                'StringValue': opts.sqs
            },
            'total': {
                'DataType': 'Number',
                'StringValue': str(list_size)
            },
            'state_abbr': {
                'DataType': 'String',
                'StringValue': 'CA'
            },
            'state_name': {
                'DataType': 'String',
                'StringValue': 'California'
            },
            'esm_uuid': {
                'DataType': 'String',
                'StringValue': esm_uuid
            }
        },
        DelaySeconds=10
    )


def send_to_api(opts, output_list, list_size):
    http = urllib3.PoolManager()
    auth = {'Authorization': f"Bearer {opts.token}"}
    in_reply_to = None
    for i, output in enumerate(output_list):
        params = {'status': f"{output}\n#Warn #Act #WarnAct #CA #California ({i+1}/{list_size})"}
        if in_reply_to:
            # Sleep a little, to avoid offending rate limiting rules?
            time.sleep(10)
            params['in_reply_to_id'] = in_reply_to
        result = http.request('POST', f"https://{opts.server}/api/v1/statuses",
                              headers=auth,
                              fiels=params)
        if result.status == 200:
            print(f"Posted {i+1}/{list_size} successfully.")
            in_reply_to = result.json()['id']
        else:
            print(f"Posting failed: {result.status}")
            print(result.data)
            sys.exit(1)


def do_update(opts, o_sheet, headers, offset, useful_columns):
    fname = opts.summary
    csv_headers = None
    rows = []
    dupes = {}

    # During the same run where we add the header, we need to ignore
    # it for the duplicate checking because the new spreadsheet added
    # it retroactively for rows we've already processed and posted
    # previously.
    #
    # Afterward we do add that new information to the existing rows so
    # that the next time, when the new column does exist already, our
    # duplicate checking doesn't go haywire in spite of it all.
    ri_added = False

    try:
        with open(fname, newline='', encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if csv_headers is None:
                    ri_seen = False
                    csv_headers = row
                    for i, col in enumerate(row):
                        if col == "Received Date":
                            # Migrate from Received Date to Processed Date
                            row[i] = "Processed Date"
                        elif col == "Related Industry":
                            ri_seen = True
                    # Migration step, because the "Related Industry"
                    # column did not exist when I started tracking
                    # this in 2023.
                    if "Related Industry" in headers and not ri_seen:
                        print("Adding 'Related Industry' to CSV headers.")
                        csv_headers.append("Related Industry")
                        ri_added = True
                else:
                    rows.append(row)
                    hashed = hashlib.sha256()
                    for col in row:
                        hashed.update(col.encode("utf-8"))
                    dupes[hashed.digest()] = len(rows) - 1
    except IOError:
        print(f"File {csv} did not exist yet.")
    if csv_headers is None:
        # No headers yet? Copy the ones from the spreadsheet
        csv_headers = []
        for col in range(useful_columns):
            o_cell = o_sheet.cell(row=1+offset, column=col+1)
            header = o_cell.value
            header = header.replace("\n", " ")
            header = header.replace("/ ", "/")
            header = header.replace("  ", " ")
            csv_headers.append(header)
        print(f"<{csv_headers}>")
    else:
        if len(csv_headers) != useful_columns:
            print("Number of columns mismatch between existing data and new data.")
            sys.exit(1)
        for header in csv_headers:
            if header not in headers:
                print(f"Header '{header}' not present in new data.")
                sys.exit(1)

    newrows = []
    dupes_total = 0
    updates_total = 0
    merged_total = 0
    for row in range(o_sheet.max_row - 1 - offset):
        newrow = []
        hashed = hashlib.sha256()
        # Pre-check, to abort if we've reached the end of useful data in the sheet
        county = o_sheet.cell(row=row+2+offset, column=headers["County/Parish"]).value
        if not county or county == "Report Summary":
            break
        for header in csv_headers:
            value = o_sheet.cell(row=row+2+offset, column=headers[header]).value
            if isinstance(value, datetime.datetime):
                value = value.strftime("%Y-%m-%d")
            else:
                value = str(value)
            newrow.append(value)
            if not ri_added or header != "Related Industry":
                hashed.update(value.encode("utf-8"))
        digest = hashed.digest()
        if digest in dupes:
            if ri_added:
                # Backfill this entry in the original CSV row, so that
                # future duplicate checking will work
                value = o_sheet.cell(row=row+2+offset, column=headers["Related Industry"]).value
                rows[dupes[digest]].append(str(value))
                merged_total += 1
            dupes_total += 1
        else:
            if opts.debug:
                print(f"New row: {newrow}")
            rows.append(newrow)
            dupes[digest] = len(rows) - 1
            newrows.append(newrow)
            updates_total += 1
    if opts.debug:
        print(f"{dupes_total} existing rows, {merged_total} merged rows, {updates_total} new rows.")
    if updates_total > 0:
        tmp_fname = f"{fname}.{os.getpid()}"
        if opts.debug:
            print(f"Creating temporary file {tmp_fname}.")
        with open(tmp_fname, 'w', newline='', encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(csv_headers)
            for row in rows:
                writer.writerow(row)
            csvfile.close()
        if opts.debug:
            print(f"Renaming {tmp_fname} to {fname}")
        os.rename(tmp_fname, fname)
    if opts.verbose:
        if len(newrows) > 0:
            print("New entries:")
            print("\n".join(dump_entries(newrows, csv_headers)))
        else:
            print("No new entries.")
    if len(newrows) > 0 and opts.post:
        output_list = dump_entries(newrows, csv_headers, False)
        list_size = len(output_list)
        if opts.sqs:
            send_to_sqs(opts, output_list, list_size)
        else:
            send_to_api(opts, output_list, list_size)


# Call from EventBridge, to replace this cron job:
#
#  process_report.py --fetch --debug
#  process_report.py --verbose --update --post --server <server> --token <token>
#
# We know that event & lambda_context are unused; '_' prefix avoids complaint
def report_handler(_event, _lambda_context):
    s3_name = os.environ['S3_NAME']
    sqs_url = os.environ['SQS_URL']

    esm_uuid = os.environ['ESM_UUID']

    # Testing:
    aws_lambda = boto3.client("lambda")
    result = aws_lambda.get_event_source_mapping(UUID=esm_uuid)
    print(f"result = {result}")
    opts = parse_options()

    os.chdir("/tmp")
    # Download the most recent spreadsheet and CSV file from S3
    # This should be okay as long as we stay under 512MB in /tmp
    #
    # https://docs.aws.amazon.com/lambda/latest/dg/gettingstarted-limits.html
    s3_resource = boto3.resource("s3")
    bucket = s3_resource.Bucket(s3_name)
    bucket.download_file('CA/warn_report.xlsx', 'warn_report.xlsx')
    bucket.download_file('CA/summary.csv', 'summary.csv')

    # process_report.py --fetch --debug
    opts.debug = True
    do_fetch(opts)

    # process_report.py --update --sqs <sqs_url>
    opts.debug = False
    opts.verbose = True
    opts.post = True
    opts.sqs = sqs_url
    o_sheet, headers, offset, useful_columns = load_report(opts)
    do_update(opts, o_sheet, headers, offset, useful_columns)

    # Upload the (potentially) updated spreadsheet and CSV file to S3
    bucket.upload_file('warn_report.xlsx', 'CA/warn_report.xlsx')
    bucket.upload_file('summary.csv', 'CA/summary.csv')


def main():
    opts = parse_options()

    if opts.dump:
        o_sheet, headers, offset, _ = load_report(opts)
        return do_dump(o_sheet, headers, offset)
    if opts.fetch:
        return do_fetch(opts)
    if opts.search:
        return do_search(opts)
    if opts.update:
        o_sheet, headers, offset, useful_columns = load_report(opts)
        return do_update(opts, o_sheet, headers, offset, useful_columns)
    print("Not Yet Implemented.")
    return False


if __name__ == "__main__":
    main()
