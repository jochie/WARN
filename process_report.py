#!/usr/bin/env python3
# -*- mode: python; tab-width: 4; indent-tabs-mode: nil -*- for emacs

import argparse
import csv
import datetime
import hashlib
import json
import os
import re
import sys
import time

import openpyxl
import requests

WARN_URL  = 'https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report.xlsx'
XLSX_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def parse_options():
    parser = argparse.ArgumentParser(
        description="""
Reading and processing the California WARN act (Excel) spreadsheet as it can be found here:

https://edd.ca.gov/en/jobs_and_training/Layoff_Services_WARN
"""
    )
    parser.add_argument('-d', '--debug',
                        help="Enable debug output",
                        default=False,
                        action='store_true')
    parser.add_argument('-v', '--verbose',
                        help="Enable verbose output",
                        default=False,
                        action='store_true')
    parser.add_argument('-n', '--dryrun',
                        help="Request dryrun (noop) mode",
                        default=False,
                        action='store_true')
    parser.add_argument("--summary",
                        help="Specify an alternative name for 'summary.csv'",
                        default="summary.csv")
    parser.add_argument("--excel",
                        help="Specify an alternative name for 'warn_report.xlsx'",
                        default="warn_report.xlsx")
    parser.add_argument("--server",
                        help="Specify an alternative Mastodon server",
                        default="botsin.space")
    parser.add_argument("--token",
                        help="Specify the authorization token. Required for --post option.")
    parser.add_argument("--post",
                        help="Post this to Mastodon? Only use with --update.",
                        default=False,
                        action='store_true')

    # The possible actions:
    parser.add_argument('--dump',
                        help="Dump a summary of the entries in the WARN act spreadsheet",
                        default=False,
                        action='store_true')
    parser.add_argument('--fetch',
                        help="Fetch the latest version of the WARN act spreadsheet from its source",
                        default=False,
                        action='store_true')
    parser.add_argument('--update',
                        help="Update the extracted data with this WARN act spreadsheet",
                        default=False,
                        action="store_true")
    parser.add_argument('--search',
                        help="Search entries matching a company in the summary.csv file")

    opts = parser.parse_args()
    excl = 0
    if opts.dump:
        excl += 1
    if opts.fetch:
        excl += 1
    if opts.search:
        excl += 1
    if opts.update:
        excl += 1
    if excl > 1:
        print("The options --dump, --fetch, --search, and --update are mutually exclusive.")
        sys.exit(1)
    if excl == 0:
        opts.dump = True
    if opts.post and not opts.update:
        print("The option --post can only be used in combination with --update.")
        sys.exit(1)
    if opts.post and not opts.token:
        print("The option --post requires that you also use the --token option.")
        sys.exit(1)
    return opts

def load_report(opts):
    workbook = openpyxl.load_workbook(filename=opts.excel, data_only=True)

    # Get All Sheets
    a_sheet_names = workbook.sheetnames
    # print(a_sheet_names)

    o_sheet = workbook[a_sheet_names[0]]
    # print(o_sheet)
    # print(o_sheet.max_row)
    # print(o_sheet.max_column)
    headers = {}
    for col in range(o_sheet.max_column):
        o_cell = o_sheet.cell(row=1, column=col+1)
        header = o_cell.value
        header = header.replace("\n", " ")
        header = header.replace("/ ", "/")
        headers[header] = col + 1
        # print(f"(1,{col+1}) = {header}")
    if opts.debug:
        print(json.dumps(headers))

    # Sanity check a few of the expected headers:
    if "No. Of Employees" not in headers:
        print("Missing No. Of Employees column, aborting.")
        sys.exit(1)
    if "Notice Date" not in headers:
        print("Missing Notice Date column, aborting.")
        sys.exit(1)

    return o_sheet, headers

# XXX: First sort the rows by Notice Date (primary) and Company (secondary)
def dump_entries(rows, csv_headers, align=True):
    output_list = []
    output = ""
    headers = {}
    for col, header in enumerate(csv_headers):
        headers[header] = col
    notice_col = headers["Notice Date"]
    last_notice = None
    for row in rows:
        notice = row[notice_col]
        if not last_notice or last_notice != notice:
            output += f"NOTICE DATE: {notice}\n\n"
            last_notice = notice
        for col, header in enumerate(csv_headers):
            if col == notice_col:
                continue
            value  = row[col]
            header = header.replace("\n", " ")
            header = header.replace("/ ", "/")
            if align:
                output += f"  {header:16s} : {value}\n"
            else:
                output += f"{header}: {value}\n"
        output_list.append(output)
        output = ""
    return output_list

def do_dump(o_sheet, headers):
    counties = {}
    companies = {}
    for row in range(o_sheet.max_row - 3):
        county = o_sheet.cell(row=row+2, column=headers["County/Parish"]).value
        if county not in counties:
            counties[county] = 0
        counties[county] += 1
        company = o_sheet.cell(row=row+2, column=headers["Company"]).value
        if company not in companies:
            companies[company] = {}
        action = o_sheet.cell(row=row+2, column=headers["Layoff/Closure"]).value
        employees = o_sheet.cell(row=row+2, column=headers["No. Of Employees"]).value
        if action not in companies[company]:
            companies[company][action] = 0
        companies[company][action] += employees
    print(json.dumps(counties))
    print(json.dumps(companies))

def do_fetch(opts):
    result = requests.get(WARN_URL)
    if result.status_code != 200:
        print(f"Unexpected HTTP status code: {result.status_code}")
        sys.exit(1)
    if result.headers['content-type'] != XLSX_TYPE:
        print(f"Unexpect content received: {result.headers['content-type']}; {result.encoding}")
        sys.exit(1)
    fname = opts.excel
    tmp_fname = f"{fname}.{os.getpid()}"
    if opts.debug:
        print(f"Creating temporary file {tmp_fname}.")
    with open(tmp_fname, 'wb') as excel:
        excel.write(result.content)
        excel.close()
    if opts.debug:
        print(f"Renaming {tmp_fname} to {fname}")
    os.rename(tmp_fname, fname)

def do_search(opts):
    fname = opts.summary
    csv_headers = None
    rows = []

    try:
        with open(fname, newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if csv_headers is None:
                    csv_headers = row
                else:
                    rows.append(row)
    except IOError:
        print(f"File {csv} does not exist.")
        sys.exit(1)

    headers = {}
    for col, header in enumerate(csv_headers):
        headers[header] = col
    if opts.debug:
        print(f"Searching for {opts.search} among {len(rows)} rows of data.")
    rows_found = []
    for row in rows:
        company = row[headers["Company"]]
        if re.match(opts.search, company):
            rows_found.append(row)
    if len(rows_found) > 0:
        print("\n".join(dump_entries(rows_found, csv_headers)))
    else:
        print("No matching companies found.")

def do_update(opts, o_sheet, headers):
    fname = opts.summary
    csv_headers = None
    rows = []
    dupes = {}

    try:
        with open(fname, newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if csv_headers is None:
                    csv_headers = row
                else:
                    rows.append(row)
                    hashed = hashlib.sha256()
                    for col in row:
                        hashed.update(col.encode("utf-8"))
                    dupes[hashed.digest()] = True
    except IOError:
        print(f"File {csv} did not exist yet.")
    if csv_headers is None:
        # No headers yet? Copy the ones from the spreadsheet
        csv_headers = []
        for col in range(o_sheet.max_column):
            o_cell = o_sheet.cell(row=1, column=col+1)
            header = o_cell.value
            header = header.replace("\n", " ")
            header = header.replace("/ ", "/")
            csv_headers.append(header)
    else:
        if len(csv_headers) != o_sheet.max_column:
            print("Number of columns mismatch between existing data and new data.")
            sys.exit(1)
        for header in csv_headers:
            if header not in headers:
                print(f"Header '{header}' not present in new data.")
                sys.exit(1)

    newrows = []
    dupes_total = 0
    updates_total = 0
    for row in range(o_sheet.max_row - 3):
        newrow = []
        hashed = hashlib.sha256()
        for header in csv_headers:
            value = o_sheet.cell(row=row+2, column=headers[header]).value
            if isinstance(value, datetime.datetime):
                value = value.strftime("%Y-%m-%d")
            else:
                value = str(value)
            newrow.append(value)
            hashed.update(value.encode("utf-8"))
        digest = hashed.digest()
        if digest not in dupes:
            dupes[digest] = True
            if opts.debug:
                print(f"New row: {newrow}")
            rows.append(newrow)
            newrows.append(newrow)
            updates_total += 1
        else:
            dupes_total += 1
    if opts.debug:
        print(f"{dupes_total} existing rows, {updates_total} new rows.")
    if updates_total > 0:
        tmp_fname = f"{fname}.{os.getpid()}"
        if opts.debug:
            print(f"Creating temporary file {tmp_fname}.")
        with open(tmp_fname, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(csv_headers)
            for row in rows:
                writer.writerow(row)
            csvfile.close()
        if opts.debug:
            print(f"Renaming {tmp_fname} to {fname}")
        os.rename(tmp_fname, fname)
    if opts.verbose:
        if len(newrows) > 0:
            print("New entries:")
            print("\n".join(dump_entries(newrows, csv_headers)))
        else:
            print("No new entries.")
    if len(newrows) > 0 and opts.post:
        auth = {'Authorization': f"Bearer {opts.token}"}
        in_reply_to = None
        output_list = dump_entries(newrows, csv_headers, False)
        list_size = len(output_list)
        for i, output in enumerate(output_list):
            params = {'status': f"{output}\n({i+1}/{list_size})"}
            if in_reply_to:
                # Sleep a little, to avoid offending rate limiting rules?
                time.sleep(5)
                params['in_reply_to_id'] = in_reply_to
            result = requests.post(f"https://{opts.server}/api/v1/statuses",
                                   data=params, headers=auth)
            if result.status_code == 200:
                print(f"Posted {i+1}/{list_size} successfully.")
                in_reply_to = result.json()['id']
            else:
                print(f"Posting failed: {result.status_code}")
                print(result.text)
                sys.exit(1)


def main():
    opts = parse_options()

    if opts.dump:
        o_sheet, headers = load_report(opts)
        return do_dump(o_sheet, headers)
    if opts.fetch:
        return do_fetch(opts)
    if opts.search:
        return do_search(opts)
    if opts.update:
        o_sheet, headers = load_report(opts)
        return do_update(opts, o_sheet, headers)
    print("Not Yet Implemented.")
    return False

if __name__ == "__main__":
    main()
